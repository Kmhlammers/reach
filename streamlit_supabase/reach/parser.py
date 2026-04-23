import io
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

KNOWN_SUPPLIERS = [
    "Evernew Arts & Crafts Co. Ltd.",
    "JINHUA JUSHI TECHNOLOGY CO.,LTD",
    "Litbright",
    "SHAOXING HIKING CANDLE GIFTS CO.,LTD",
    "SHAOXING RUIYAO TRADING CO.,LTD",
    "TIANJIN NATIVE PRODUCE IMPORT AND EXPORT GROUP CORPORATION LIMITED",
    "KWUNGS",
    "Marvellite Aromatics Pvt. Ltd.",
    "Ningbo Jiangdong Shine Star Imp.& Exp. Co.,Ltd",
    "Ningbo Zhongning Import & Export Co. Ltd./Ninghai Haolaiyun Household Utensils Co.,Ltd",
]


def normalize(value: str) -> str:
    return str(value or "").lower().strip()


def normalize_supplier_name(value: str) -> str:
    return (
        str(value or "")
        .lower()
        .replace("&", " and ")
        .replace(".", " ")
        .replace(",", " ")
        .replace("(", " ")
        .replace(")", " ")
        .replace("/", " ")
        .replace("-", " ")
        .strip()
    )


def normalize_cas_number(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").replace(" - ", "-").replace("- ", "-").replace(" -", "-")).strip()


def is_valid_cas_number(value: str) -> bool:
    return bool(re.match(r"^\d{2,7}-\d{2}-\d$", normalize_cas_number(value)))


def to_num(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^\d.\-]", "", str(value).replace(",", "."))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def rows_to_matrix(ws) -> List[List[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def get_cell_text(values: List[List[Any]], r: int, c: int) -> str:
    try:
        v = values[r][c]
        return "" if v is None else str(v).strip()
    except Exception:
        return ""


def find_label_value(values: List[List[Any]], labels: List[str]) -> Tuple[str, str]:
    targets = [normalize(x) for x in labels]

    for r in range(len(values)):
        for c in range(len(values[r])):
            text = get_cell_text(values, r, c)
            n = normalize(text)
            if not n:
                continue

            for target in targets:
                if n == target or target in n:
                    lower_text = text.lower()
                    idx = lower_text.find(target)
                    if idx >= 0:
                        inline = text[idx + len(target) :].strip(" :-–—")
                        if inline:
                            return text, inline

                    for c2 in range(c + 1, min(c + 12, len(values[r]))):
                        right = get_cell_text(values, r, c2)
                        if right:
                            return text, right
    return "", ""


def find_total_value(values: List[List[Any]], labels: List[str]) -> Tuple[str, float]:
    targets = [normalize(x) for x in labels]

    for r in range(len(values) - 1, -1, -1):
        for c in range(len(values[r])):
            text = get_cell_text(values, r, c)
            n = normalize(text)
            if not n:
                continue
            for target in targets:
                if n == target or target in n:
                    last_number = 0.0
                    found = False
                    for c2 in range(c + 1, len(values[r])):
                        v = values[r][c2]
                        if v is None or str(v).strip() == "":
                            continue
                        last_number = to_num(v)
                        found = True
                    return text, last_number if found else 0.0
    return "", 0.0


def match_known_supplier(raw_supplier: str) -> Tuple[str, str, bool]:
    original = str(raw_supplier or "").strip()
    normalized_input = normalize_supplier_name(original)

    if not normalized_input:
        return "", "", False

    for known in KNOWN_SUPPLIERS:
        if normalize_supplier_name(known) == normalized_input:
            return original, known, True

    for known in KNOWN_SUPPLIERS:
        if normalize_supplier_name(known) in normalized_input or normalized_input in normalize_supplier_name(known):
            return original, known, True

    return original, original, False


def parse_workbook_bytes(file_bytes: bytes, payload: Dict[str, Any]) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    now_iso = datetime.now(timezone.utc).isoformat()

    source_type = payload.get("source_type", "upload")
    source_ref = payload.get("source_ref")
    file_name = payload.get("file_name", "uploaded.xlsx")

    records: List[Dict[str, Any]] = []
    issues: List[Dict[str, Any]] = []
    seen = set()

    workbook_supplier_original = ""
    workbook_supplier_matched = ""

    for ws in wb.worksheets:
        values = rows_to_matrix(ws)
        _, supplier = find_label_value(values, ["supplier name:", "supplier name"])
        if supplier:
            workbook_supplier_original, workbook_supplier_matched, _ = match_known_supplier(supplier)
            break

    for ws in wb.worksheets:
        values = rows_to_matrix(ws)
        sheet_name = ws.title

        _, substance = find_label_value(values, ["substance name (a)", "substance name"])
        _, cas = find_label_value(values, ["cas-number:", "cas number:", "cas-number", "cas number"])
        _, or_registered = find_label_value(
            values,
            [
                "is the ingredient registered by only representative company (or)?",
                "is the ingredient registered by only representative company",
                "registered by only representative company",
                "registered by or",
                "or?",
            ],
        )

        _, total_kg = find_total_value(
            values, ["total weight of substance (kg)", "total weight of substance kg", "total kg", "kg"]
        )
        _, total_tonnes = find_total_value(
            values,
            [
                "total weight of substance (tons)",
                "total weight of substance (ton)",
                "total tonnes",
                "total tons",
                "total ton",
                "tons",
            ],
        )

        cas = normalize_cas_number(cas)
        has_data = bool(substance or cas or or_registered or total_kg or total_tonnes)
        if not has_data:
            issues.append(
                {
                    "sheet": sheet_name,
                    "issue_type": "no_meaningful_data",
                    "message": "No meaningful data found on this sheet.",
                    "row_key": f"{file_name}|{sheet_name}|empty|empty",
                }
            )
            continue

        import_message = ""
        import_status = "processed"

        if not is_valid_cas_number(cas):
            if substance:
                import_message = f"Invalid CAS '{cas}'. Used substance as fallback key."
                cas = substance.strip()
                import_status = "fallback"
            else:
                issues.append(
                    {
                        "sheet": sheet_name,
                        "issue_type": "invalid_cas",
                        "message": f"Invalid CAS '{cas}' and no usable substance fallback.",
                        "row_key": f"{file_name}|{sheet_name}|invalid|empty",
                    }
                )
                continue

        row_key = f"{file_name}|{sheet_name}|{cas}|{substance}".strip()
        if row_key in seen:
            issues.append(
                {
                    "sheet": sheet_name,
                    "issue_type": "duplicate",
                    "message": "Duplicate record skipped.",
                    "row_key": row_key,
                }
            )
            continue

        seen.add(row_key)

        records.append(
            {
                "row_key": row_key,
                "sheet": sheet_name,
                "supplier_name": workbook_supplier_matched or workbook_supplier_original,
                "supplier_name_original": workbook_supplier_original,
                "cas_number": cas,
                "substance": substance,
                "or_registered": or_registered,
                "total_kg": float(total_kg),
                "total_tonnes": float(total_tonnes),
                "import_status": import_status,
                "import_message": import_message,
                "processed_at": now_iso,
            }
        )

    return {
        "ok": True,
        "file": {
            "file_name": file_name,
            "source_type": source_type,
            "source_ref": source_ref,
            "sheet_count": len(wb.worksheets),
        },
        "summary": {
            "records_found": len(records),
            "issues_count": len(issues),
        },
        "records": records,
        "issues": issues,
    }

