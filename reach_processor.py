import io
import json
import logging
import os
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote, urlparse

import requests
from openpyxl import load_workbook

GRAPH_BASE = "https://graph.microsoft.com/v1.0"

KNOWN_SUPPLIERS = [
    "Evernew Arts & Crafts Co. Ltd.",
    "JINHUA JUSHI TECHNOLOGY CO.,LTD",
    "Litbright",
    "SHAOXING HIKING CANDLE GIFTS CO.,LTD",
    "SHAOXING RUIYAO TRADING CO.,LTD",
    "TIANJIN NATIVE PRODUCE IMPORT AND EXPORT GROUP CORPORATION LIMITED",
    "KWUNGS",
    "Marvellite Aromatics Pvt. Ltd.",
    "Ningbo Jiangdong Shine Star Imp.& Exp. Co.,Ltd",
    "Ningbo Zhongning Import & Export Co. Ltd./Ninghai Haolaiyun Household Utensils Co.,Ltd",
]


class GraphClient:
    def __init__(self) -> None:
        self.tenant_id = self._required_env("AZURE_TENANT_ID")
        self.client_id = self._required_env("AZURE_CLIENT_ID")
        self.client_secret = self._required_env("AZURE_CLIENT_SECRET")
        self.summary_list_name = os.getenv("REACH_SUMMARY_LIST_NAME", "REACH_Summary")
        self.totals_list_name = os.getenv("REACH_TOTALS_LIST_NAME", "Reach total of substances")
        self.timeout = int(os.getenv("GRAPH_TIMEOUT_SECONDS", "60"))
        self._token: Optional[str] = None

    @staticmethod
    def _required_env(name: str) -> str:
        value = os.getenv(name)
        if not value:
            raise RuntimeError(f"Missing required environment variable: {name}")
        return value

    def _acquire_token(self) -> str:
        if self._token:
            return self._token

        token_url = f"https://login.microsoftonline.com/{self.tenant_id}/oauth2/v2.0/token"
        response = requests.post(
            token_url,
            data={
                "client_id": self.client_id,
                "client_secret": self.client_secret,
                "scope": "https://graph.microsoft.com/.default",
                "grant_type": "client_credentials",
            },
            timeout=self.timeout,
        )
        response.raise_for_status()
        data = response.json()
        self._token = data["access_token"]
        return self._token

    def _headers(self) -> Dict[str, str]:
        return {
            "Authorization": f"Bearer {self._acquire_token()}",
            "Accept": "application/json",
        }

    def get_json(self, url: str, **kwargs) -> Dict[str, Any]:
        response = requests.get(url, headers=self._headers(), timeout=self.timeout, **kwargs)
        self._raise_for_graph_error(response)
        return response.json()

    def get_bytes(self, url: str, **kwargs) -> bytes:
        response = requests.get(url, headers=self._headers(), timeout=self.timeout, **kwargs)
        self._raise_for_graph_error(response)
        return response.content

    def post_json(self, url: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        headers = {**self._headers(), "Content-Type": "application/json"}
        response = requests.post(url, headers=headers, json=payload, timeout=self.timeout)
        self._raise_for_graph_error(response)
        return response.json() if response.content else {}

    def patch_json(self, url: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        headers = {**self._headers(), "Content-Type": "application/json"}
        response = requests.patch(url, headers=headers, json=payload, timeout=self.timeout)
        self._raise_for_graph_error(response)
        return response.json() if response.content else {}

    def delete(self, url: str) -> None:
        response = requests.delete(url, headers=self._headers(), timeout=self.timeout)
        self._raise_for_graph_error(response)

    @staticmethod
    def _raise_for_graph_error(response: requests.Response) -> None:
        if response.ok:
            return
        try:
            payload = response.json()
        except Exception:
            payload = {"raw": response.text}
        raise RuntimeError(f"Graph API error {response.status_code}: {json.dumps(payload)}")

    def get_site_id_from_url(self, site_url: str) -> str:
        parsed = urlparse(site_url)
        hostname = parsed.netloc
        site_path = parsed.path.rstrip("/")
        if not hostname or not site_path:
            raise RuntimeError(f"Invalid site_url: {site_url}")

        url = f"{GRAPH_BASE}/sites/{hostname}:{site_path}"
        data = self.get_json(url)
        return data["id"]

    def get_list_id_by_name(self, site_id: str, list_name: str) -> str:
        url = f"{GRAPH_BASE}/sites/{site_id}/lists/{quote(list_name)}"
        data = self.get_json(url)
        return data["id"]

    def download_file_bytes(self, site_id: str, file_path: str) -> bytes:
        normalized = file_path.strip()
        if not normalized.startswith("/"):
            normalized = "/" + normalized

        encoded_path = quote(normalized, safe="/-_.() ")
        encoded_path = encoded_path.replace(" ", "%20")
        url = f"{GRAPH_BASE}/sites/{site_id}/drive/root:{encoded_path}:/content"
        return self.get_bytes(url, allow_redirects=True)

    def list_items(
        self,
        site_id: str,
        list_id: str,
        *,
        expand_fields: bool = True,
        select_fields: Optional[List[str]] = None,
        filter_query: Optional[str] = None,
        top: int = 200,
    ) -> List[Dict[str, Any]]:
        params = {"$top": str(top)}
        if expand_fields:
            if select_fields:
                params["$expand"] = f"fields($select={','.join(select_fields)})"
            else:
                params["$expand"] = "fields"
        if filter_query:
            params["$filter"] = filter_query

        url = f"{GRAPH_BASE}/sites/{site_id}/lists/{list_id}/items"
        items: List[Dict[str, Any]] = []

        while url:
            data = self.get_json(url, params=params if url.endswith("/items") else None)
            items.extend(data.get("value", []))
            url = data.get("@odata.nextLink")
            params = None

        return items

    def create_list_item(self, site_id: str, list_id: str, fields: Dict[str, Any]) -> Dict[str, Any]:
        url = f"{GRAPH_BASE}/sites/{site_id}/lists/{list_id}/items"
        return self.post_json(url, {"fields": fields})

    def update_list_item_fields(self, site_id: str, list_id: str, item_id: str, fields: Dict[str, Any]) -> Dict[str, Any]:
        url = f"{GRAPH_BASE}/sites/{site_id}/lists/{list_id}/items/{item_id}/fields"
        return self.patch_json(url, fields)

    def delete_list_item(self, site_id: str, list_id: str, item_id: str) -> None:
        url = f"{GRAPH_BASE}/sites/{site_id}/lists/{list_id}/items/{item_id}"
        self.delete(url)


# ---------- workbook parsing ----------

def normalize(value: str) -> str:
    return str(value or "").lower().strip()


def normalize_supplier_name(value: str) -> str:
    return (
        str(value or "")
        .lower()
        .replace("&", " and ")
        .replace(".", " ")
        .replace(",", " ")
        .replace("(", " ")
        .replace(")", " ")
        .replace("/", " ")
        .replace("-", " ")
        .strip()
    )


def normalize_cas_number(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").replace(" - ", "-").replace("- ", "-").replace(" -", "-")).strip()


def is_valid_cas_number(value: str) -> bool:
    return bool(re.match(r"^\d{2,7}-\d{2}-\d$", normalize_cas_number(value)))


def to_num(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^\d.\-]", "", str(value).replace(",", "."))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def rows_to_matrix(ws) -> List[List[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def get_cell_text(values: List[List[Any]], r: int, c: int) -> str:
    try:
        v = values[r][c]
        return "" if v is None else str(v).strip()
    except Exception:
        return ""


def find_label_value(values: List[List[Any]], labels: List[str]) -> Tuple[str, str]:
    targets = [normalize(x) for x in labels]

    for r in range(len(values)):
        for c in range(len(values[r])):
            text = get_cell_text(values, r, c)
            n = normalize(text)
            if not n:
                continue

            for target in targets:
                if n == target or target in n:
                    lower_text = text.lower()
                    idx = lower_text.find(target)
                    if idx >= 0:
                        inline = text[idx + len(target):].strip(" :-–—")
                        if inline:
                            return text, inline

                    for c2 in range(c + 1, min(c + 12, len(values[r]))):
                        right = get_cell_text(values, r, c2)
                        if right:
                            return text, right
    return "", ""


def find_total_value(values: List[List[Any]], labels: List[str]) -> Tuple[str, float]:
    targets = [normalize(x) for x in labels]

    for r in range(len(values) - 1, -1, -1):
        for c in range(len(values[r])):
            text = get_cell_text(values, r, c)
            n = normalize(text)
            if not n:
                continue
            for target in targets:
                if n == target or target in n:
                    last_number = 0.0
                    found = False
                    for c2 in range(c + 1, len(values[r])):
                        v = values[r][c2]
                        if v is None or str(v).strip() == "":
                            continue
                        last_number = to_num(v)
                        found = True
                    return text, last_number if found else 0.0
    return "", 0.0


def match_known_supplier(raw_supplier: str) -> Tuple[str, str, bool]:
    original = str(raw_supplier or "").strip()
    normalized_input = normalize_supplier_name(original)

    if not normalized_input:
        return "", "", False

    for known in KNOWN_SUPPLIERS:
        if normalize_supplier_name(known) == normalized_input:
            return original, known, True

    for known in KNOWN_SUPPLIERS:
        if normalize_supplier_name(known) in normalized_input or normalized_input in normalize_supplier_name(known):
            return original, known, True

    return original, original, False


def parse_workbook_bytes(file_bytes: bytes, payload: Dict[str, Any]) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    now_iso = datetime.now(timezone.utc).isoformat()
    today = datetime.now(timezone.utc).date().isoformat()

    file_id = str(payload["file_id"])
    file_name = payload["file_name"]
    file_link = payload["file_link"]
    file_path = payload["file_path"]
    source_type = payload["source_type"]

    records: List[Dict[str, Any]] = []
    issues: List[Dict[str, Any]] = []
    seen = set()

    workbook_supplier_original = ""
    workbook_supplier_matched = ""

    for ws in wb.worksheets:
        values = rows_to_matrix(ws)
        _, supplier = find_label_value(values, ["supplier name:", "supplier name"])
        if supplier:
            workbook_supplier_original, workbook_supplier_matched, _ = match_known_supplier(supplier)
            break

    for ws in wb.worksheets:
        values = rows_to_matrix(ws)
        sheet_name = ws.title

        _, substance = find_label_value(values, ["substance name (a)", "substance name"])
        _, cas = find_label_value(values, ["cas-number:", "cas number:", "cas-number", "cas number"])
        _, or_registered = find_label_value(
            values,
            [
                "is the ingredient registered by only representative company (or)?",
                "is the ingredient registered by only representative company",
                "registered by only representative company",
                "registered by or",
                "or?",
            ],
        )

        _, total_kg = find_total_value(values, ["total weight of substance (kg)", "total weight of substance kg", "total kg", "kg"])
        _, total_tonnes = find_total_value(values, ["total weight of substance (tons)", "total weight of substance (ton)", "total tonnes", "total tons", "total ton", "tons"])

        cas = normalize_cas_number(cas)
        has_data = bool(substance or cas or or_registered or total_kg or total_tonnes)
        if not has_data:
            issues.append(
                {
                    "sheet": sheet_name,
                    "issue_type": "no_meaningful_data",
                    "message": "No meaningful data found on this sheet.",
                    "row_key": f"{file_id}|{sheet_name}|empty|empty",
                }
            )
            continue

        import_message = ""
        import_status = "processed"

        if not is_valid_cas_number(cas):
            if substance:
                import_message = f"Invalid CAS '{cas}'. Used substance as fallback key."
                cas = substance.strip()
                import_status = "fallback"
            else:
                issues.append(
                    {
                        "sheet": sheet_name,
                        "issue_type": "invalid_cas",
                        "message": f"Invalid CAS '{cas}' and no usable substance fallback.",
                        "row_key": f"{file_id}|{sheet_name}|invalid|empty",
                    }
                )
                continue

        row_key = f"{file_id}|{sheet_name}|{cas}|{substance}".strip()
        if row_key in seen:
            issues.append(
                {
                    "sheet": sheet_name,
                    "issue_type": "duplicate",
                    "message": "Duplicate record skipped.",
                    "row_key": row_key,
                }
            )
            continue

        seen.add(row_key)

        records.append(
            {
                "Title": row_key,
                "FileId": file_id,
                "FileName": file_name,
                "RowKey": row_key,
                "Suppliername": workbook_supplier_matched or workbook_supplier_original,
                "Suppliernameoriginal": workbook_supplier_original,
                "Casnumber": cas,
                "Substance": substance,
                "ORregistred": or_registered,
                "Totalkg": float(total_kg),
                "Totalkgcurrent": float(total_kg),
                "Totaltonnes": float(total_tonnes),
                "DATE": today,
                "Filelink": file_link,
                "Documentfilepath": file_path,
                "Sheet": sheet_name,
                "Processed": 1,
                "Processedat": now_iso,
                "Sourcetype": source_type,
                "Importstatus": import_status,
                "Importmessage": import_message,
                "Sendcertificate": False,
                "Mailopzoeken": False,
            }
        )

    return {
        "ok": True,
        "file": {
            "file_id": file_id,
            "file_name": file_name,
            "sheet_count": len(wb.worksheets),
            "source_type": source_type,
        },
        "summary": {
            "records_found": len(records),
            "records_skipped": len(issues),
            "issues_count": len(issues),
        },
        "records": records,
        "issues": issues,
    }


# ---------- SharePoint write logic ----------

def _boolish(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "ja"}


def sanitize_field_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value)


def get_existing_items_by_file_id(graph: GraphClient, site_id: str, list_id: str, file_id: str) -> List[Dict[str, Any]]:
    filter_query = f"fields/FileId eq '{file_id.replace("'", "''")}'"
    return graph.list_items(
        site_id,
        list_id,
        expand_fields=True,
        select_fields=["Title", "FileId", "Processed"],
        filter_query=filter_query,
    )


def sync_summary_items(
    graph: GraphClient,
    site_id: str,
    summary_list_id: str,
    file_id: str,
    new_records: List[Dict[str, Any]],
) -> Dict[str, int]:
    existing_items = get_existing_items_by_file_id(graph, site_id, summary_list_id, file_id)
    existing_by_title = {item.get("fields", {}).get("Title"): item for item in existing_items}
    new_by_title = {record["Title"]: record for record in new_records}

    created = 0
    updated = 0
    deleted = 0

    for title, record in new_by_title.items():
        cleaned_fields = {k: sanitize_field_value(v) for k, v in record.items() if v is not None}
        existing = existing_by_title.get(title)
        if existing:
            graph.update_list_item_fields(site_id, summary_list_id, existing["id"], cleaned_fields)
            updated += 1
        else:
            graph.create_list_item(site_id, summary_list_id, cleaned_fields)
            created += 1

    for title, item in existing_by_title.items():
        if title not in new_by_title:
            graph.delete_list_item(site_id, summary_list_id, item["id"])
            deleted += 1

    return {"created": created, "updated": updated, "deleted": deleted}


def aggregate_totals_from_summary(graph: GraphClient, site_id: str, summary_list_id: str) -> Dict[str, Dict[str, Any]]:
    items = graph.list_items(
        site_id,
        summary_list_id,
        expand_fields=True,
        select_fields=["Casnumber", "Substance", "Totalkg", "Totaltonnes", "Processed"],
        top=500,
    )

    totals: Dict[str, Dict[str, Any]] = {}
    for item in items:
        fields = item.get("fields", {})
        processed = fields.get("Processed")
        if str(processed) not in {"1", "1.0"} and not _boolish(processed):
            continue

        cas = str(fields.get("Casnumber") or "").strip()
        substance = str(fields.get("Substance") or "").strip()
        if not cas:
            continue

        bucket = totals.setdefault(
            cas,
            {
                "Title": cas,
                "Casnumber": cas,
                "Substance": substance,
                "Totalkg": 0.0,
                "Totaltonnes": 0.0,
            },
        )
        if not bucket.get("Substance") and substance:
            bucket["Substance"] = substance

        bucket["Totalkg"] += to_num(fields.get("Totalkg"))
        bucket["Totaltonnes"] += to_num(fields.get("Totaltonnes"))

    for bucket in totals.values():
        bucket["Totalkg"] = round(bucket["Totalkg"], 6)
        bucket["Totaltonnes"] = round(bucket["Totaltonnes"], 6)

    return totals


def sync_totals_list(graph: GraphClient, site_id: str, totals_list_id: str, totals_map: Dict[str, Dict[str, Any]]) -> Dict[str, int]:
    existing_items = graph.list_items(
        site_id,
        totals_list_id,
        expand_fields=True,
        select_fields=["Title", "Casnumber", "Totalkg", "Totaltonnes"],
        top=500,
    )
    existing_by_title = {item.get("fields", {}).get("Title"): item for item in existing_items}

    created = 0
    updated = 0
    deleted = 0

    for title, total_record in totals_map.items():
        cleaned_fields = {k: sanitize_field_value(v) for k, v in total_record.items() if v is not None}
        existing = existing_by_title.get(title)
        if existing:
            graph.update_list_item_fields(site_id, totals_list_id, existing["id"], cleaned_fields)
            updated += 1
        else:
            graph.create_list_item(site_id, totals_list_id, cleaned_fields)
            created += 1

    for title, item in existing_by_title.items():
        if title not in totals_map:
            graph.delete_list_item(site_id, totals_list_id, item["id"])
            deleted += 1

    return {"created": created, "updated": updated, "deleted": deleted}


def process_reach_workbook(payload: Dict[str, Any]) -> Dict[str, Any]:
    graph = GraphClient()
    site_id = graph.get_site_id_from_url(payload["site_url"])
    summary_list_id = graph.get_list_id_by_name(site_id, graph.summary_list_name)
    totals_list_id = graph.get_list_id_by_name(site_id, graph.totals_list_name)

    file_bytes = graph.download_file_bytes(site_id, payload["file_path"])
    parsed = parse_workbook_bytes(file_bytes, payload)

    summary_sync = sync_summary_items(
        graph=graph,
        site_id=site_id,
        summary_list_id=summary_list_id,
        file_id=str(payload["file_id"]),
        new_records=parsed["records"],
    )

    totals_map = aggregate_totals_from_summary(graph, site_id, summary_list_id)
    totals_sync = sync_totals_list(graph, site_id, totals_list_id, totals_map)

    parsed["summary"]["summary_items_created"] = summary_sync["created"]
    parsed["summary"]["summary_items_updated"] = summary_sync["updated"]
    parsed["summary"]["summary_items_deleted"] = summary_sync["deleted"]
    parsed["summary"]["totals_items_created"] = totals_sync["created"]
    parsed["summary"]["totals_items_updated"] = totals_sync["updated"]
    parsed["summary"]["totals_items_deleted"] = totals_sync["deleted"]
    parsed["sharepoint"] = {
        "site_id": site_id,
        "summary_list_id": summary_list_id,
        "totals_list_id": totals_list_id,
    }
    return parsed
